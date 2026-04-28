#!/usr/bin/env python3
"""
export_excel.py - Excel 报告导出模块

将复现结果导出为 Excel 文件，便于对比分析。

功能:
- 导出指标对比表（论文目标 vs 实际结果）
- 支持多指标对比
- 自动计算差距和评估状态

用法:
    python export_excel.py --evaluation <JSON> --output <XLSX>
    python export_excel.py --all --inputs_dir <DIR> --output_dir <DIR>
"""

import argparse
import json
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional


def install_required_packages():
    """检查并安装必要的包"""
    try:
        import openpyxl
        return True
    except ImportError:
        print("正在安装 openpyxl...")
        result = os.system("pip install openpyxl -q")
        return result == 0


def cm2score_metrics(evaluation_data: Dict[str, Any]) -> Dict[str, float]:
    """
    从评测数据中提取 cm2score 指标

    Args:
        evaluation_data: step6_evaluation.json 数据

    Returns:
        包含指标的字典
    """
    metrics = {}

    # 从 gap_analysis 提取
    for gap in evaluation_data.get("gap_analysis", []):
        metric_name = gap.get("metric_name", "")
        achieved = gap.get("achieved")
        target = gap.get("target")

        if achieved is not None:
            metrics[f"{metric_name}_achieved"] = achieved
        if target is not None:
            metrics[f"{metric_name}_target"] = target

    # 从 achieved_metrics 提取
    for metric in evaluation_data.get("achieved_metrics", {}).get("metrics", []):
        name = metric.get("metric_name", "")
        value = metric.get("achieved_value")
        if value is not None and f"{name}_achieved" not in metrics:
            metrics[f"{name}_achieved"] = value

    # 从 target_metrics 提取
    for metric in evaluation_data.get("target_metrics", {}).get("metrics", []):
        name = metric.get("metric_name", "")
        value = metric.get("target_value")
        if value is not None and f"{name}_target" not in metrics:
            metrics[f"{name}_target"] = value

    return metrics


def export_metrics_comparison(evaluation_path: str, output_path: str,
                             paper_info: Optional[Dict] = None) -> bool:
    """
    导出指标对比表到 Excel

    Args:
        evaluation_path: step6_evaluation.json 路径
        output_path: 输出 Excel 路径
        paper_info: 论文信息（可选）

    Returns:
        是否成功
    """
    if not install_required_packages():
        print("错误: 无法安装 openpyxl")
        return False

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("错误: openpyxl 未安装")
        return False

    # 读取评测数据
    with open(evaluation_path, "r", encoding="utf-8") as f:
        evaluation_data = json.load(f)

    # 创建工作簿
    wb = Workbook()

    # ===== Sheet 1: 指标对比 =====
    ws_metrics = wb.active
    ws_metrics.title = "指标对比"

    # 样式定义
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    exceeded_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    acceptable_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    warning_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 标题
    ws_metrics.merge_cells('A1:F1')
    ws_metrics['A1'] = "遥感变化检测论文复现 - 指标对比表"
    ws_metrics['A1'].font = Font(bold=True, size=14)
    ws_metrics['A1'].alignment = Alignment(horizontal='center')

    # 基本信息
    ws_metrics['A3'] = "论文信息"
    ws_metrics['A3'].font = Font(bold=True)

    if paper_info:
        ws_metrics['A4'] = "标题:"
        ws_metrics['B4'] = paper_info.get("title", "未知")
        ws_metrics['A5'] = "作者:"
        ws_metrics['B5'] = ", ".join(paper_info.get("authors", []))
        ws_metrics['A6'] = "年份:"
        ws_metrics['B6'] = paper_info.get("year", "未知")

    ws_metrics['A7'] = "复现日期:"
    ws_metrics['B7'] = datetime.now().strftime("%Y-%m-%d %H:%M")

    # 表头
    headers = ["指标名称", "论文目标", "实际结果", "绝对差距", "相对差距(%)", "评估"]
    header_row = 10
    for col, header in enumerate(headers, 1):
        cell = ws_metrics.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # 填充数据
    gap_analysis = evaluation_data.get("gap_analysis", [])
    row = header_row + 1

    for gap in gap_analysis:
        metric_name = gap.get("metric_name", "")
        target = gap.get("target")
        achieved = gap.get("achieved")
        gap_val = gap.get("gap")
        gap_pct = gap.get("gap_percentage")
        assessment = gap.get("assessment", "unknown")

        # 指标名称
        ws_metrics.cell(row=row, column=1, value=metric_name).border = thin_border
        ws_metrics.cell(row=row, column=1).alignment = Alignment(horizontal='center')

        # 论文目标
        target_cell = ws_metrics.cell(row=row, column=2, value=target if target else "N/A")
        target_cell.border = thin_border
        target_cell.number_format = '0.0000'
        target_cell.alignment = Alignment(horizontal='center')

        # 实际结果
        achieved_cell = ws_metrics.cell(row=row, column=3, value=achieved if achieved is not None else "N/A")
        achieved_cell.border = thin_border
        achieved_cell.number_format = '0.0000'
        achieved_cell.alignment = Alignment(horizontal='center')

        # 绝对差距
        gap_cell = ws_metrics.cell(row=row, column=4, value=gap_val if gap_val is not None else "N/A")
        gap_cell.border = thin_border
        gap_cell.number_format = '+0.0000;-0.0000'
        gap_cell.alignment = Alignment(horizontal='center')

        # 相对差距
        pct_cell = ws_metrics.cell(row=row, column=5, value=gap_pct / 100 if gap_pct else "N/A")
        pct_cell.border = thin_border
        pct_cell.number_format = '+0.00%;-0.00%'
        pct_cell.alignment = Alignment(horizontal='center')

        # 评估
        assessment_text = {
            "exceeded": "✅ 超过目标",
            "acceptable": "✅ 可接受",
            "mild_deviation": "⚠️ 轻微偏差",
            "significant_deviation": "⚠️ 显著偏差",
            "critical": "❌ 严重偏差",
        }.get(assessment, assessment)
        assessment_cell = ws_metrics.cell(row=row, column=6, value=assessment_text)
        assessment_cell.border = thin_border
        assessment_cell.alignment = Alignment(horizontal='center')

        # 根据评估设置填充色
        if assessment == "exceeded":
            for col in range(1, 7):
                ws_metrics.cell(row=row, column=col).fill = exceeded_fill
        elif assessment == "acceptable":
            for col in range(1, 7):
                ws_metrics.cell(row=row, column=col).fill = acceptable_fill
        elif assessment in ["significant_deviation", "critical"]:
            for col in range(1, 7):
                ws_metrics.cell(row=row, column=col).fill = warning_fill

        row += 1

    # 设置列宽
    ws_metrics.column_dimensions['A'].width = 15
    ws_metrics.column_dimensions['B'].width = 15
    ws_metrics.column_dimensions['C'].width = 15
    ws_metrics.column_dimensions['D'].width = 15
    ws_metrics.column_dimensions['E'].width = 15
    ws_metrics.column_dimensions['F'].width = 20

    # ===== Sheet 2: 详细结果 =====
    ws_detail = wb.create_sheet(title="详细结果")

    # 复现状态
    status = evaluation_data.get("reproduction_status", {})
    ws_detail['A1'] = "复现状态"
    ws_detail['A1'].font = Font(bold=True, size=12)

    status_text = {
        "fully_reproduced": "✅ 完全复现",
        "partially_reproduced": "⚠️ 部分复现",
        "failed_reproduction": "❌ 复现失败",
    }.get(status.get("overall", ""), status.get("overall", "未知"))

    ws_detail['A2'] = "整体状态:"
    ws_detail['B2'] = status_text
    ws_detail['A3'] = "代码可运行:"
    ws_detail['B3'] = "是" if status.get("code_runs") else "否"
    ws_detail['A4'] = "数据可用:"
    ws_detail['B4'] = "是" if status.get("data_available") else "否"
    ws_detail['A5'] = "指标匹配:"
    ws_detail['B5'] = "是" if status.get("metrics_match") else "否"

    # 问题列表
    issues = status.get("issues", [])
    if issues:
        ws_detail['A7'] = "问题列表"
        ws_detail['A7'].font = Font(bold=True)
        for i, issue in enumerate(issues, 8):
            ws_detail[f'A{i}'] = f"- {issue}"

    # 指标计算方法
    ws_detail['A12'] = "指标计算方法"
    ws_detail['A12'].font = Font(bold=True)
    ws_detail['A13'] = f"方法: {evaluation_data.get('metrics_calculation_method', 'cm2score')}"

    # ===== Sheet 3: 配置信息 =====
    ws_config = wb.create_sheet(title="配置信息")

    if paper_info:
        ws_config['A1'] = "论文配置"
        ws_config['A1'].font = Font(bold=True, size=12)

        model_info = paper_info.get("model_info", {})
        training_config = paper_info.get("training_config", {})

        config_items = [
            ("模型名称", model_info.get("model_name", "未知")),
            ("架构", model_info.get("architecture", "未知")),
            ("Backbone", model_info.get("backbone", "未知")),
            ("输入尺寸", str(model_info.get("input_size", []))),
            ("Epochs", training_config.get("epochs", "未知")),
            ("Batch Size", training_config.get("batch_size", "未知")),
            ("优化器", training_config.get("optimizer", "未知")),
            ("学习率", training_config.get("learning_rate", "未知")),
        ]

        for i, (key, value) in enumerate(config_items, 2):
            ws_config[f'A{i}'] = key
            ws_config[f'B{i}'] = str(value)

    # 训练环境
    ws_config['A12'] = "训练环境"
    ws_config['A12'].font = Font(bold=True, size=12)

    training = evaluation_data.get("training_config", {})
    env_items = [
        ("训练模式", training.get("mode", "未知")),
        ("Batch Size", training.get("batch_size", "未知")),
        ("学习率", training.get("learning_rate", "未知")),
        ("优化器", training.get("optimizer", "未知")),
    ]

    for i, (key, value) in enumerate(env_items, 13):
        ws_config[f'A{i}'] = key
        ws_config[f'B{i}'] = str(value)

    # 设置列宽
    ws_config.column_dimensions['A'].width = 20
    ws_config.column_dimensions['B'].width = 30

    # 保存
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    print(f"✅ Excel 报告已导出: {output_path}")
    return True


def main():
    parser = argparse.ArgumentParser(description="Excel 报告导出")
    parser.add_argument("--evaluation", "-e", help="step6_evaluation.json 路径")
    parser.add_argument("--paper", "-p", help="step1_paper_parse.json 路径（可选）")
    parser.add_argument("--output", "-o", help="输出 Excel 路径")
    parser.add_argument("--inputs_dir", "-i", help="输入目录（包含所有 JSON）")
    parser.add_argument("--output_dir", "-d", help="输出目录")

    args = parser.parse_args()

    if args.inputs_dir and args.output_dir:
        # 批量导出
        inputs_dir = Path(args.inputs_dir)

        eval_path = inputs_dir / "step6_evaluation.json"
        paper_path = inputs_dir / "step1_paper_parse.json"

        if not eval_path.exists():
            print(f"错误: 未找到 {eval_path}")
            return 1

        paper_info = None
        if paper_path.exists():
            with open(paper_path, "r", encoding="utf-8") as f:
                paper_info = json.load(f)

        output_dir = Path(args.output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / "metrics_comparison.xlsx"

        export_metrics_comparison(str(eval_path), str(output_path), paper_info)

    elif args.evaluation and args.output:
        paper_info = None
        if args.paper and os.path.exists(args.paper):
            with open(args.paper, "r", encoding="utf-8") as f:
                paper_info = json.load(f)

        export_metrics_comparison(args.evaluation, args.output, paper_info)

    else:
        parser.print_help()
        return 1

    return 0


if __name__ == "__main__":
    sys.exit(main())
